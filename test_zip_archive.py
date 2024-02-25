import zipfile
from pypdf import PdfReader
import openpyxl
from openpyxl import load_workbook
from xlrd import open_workbook
from script_os import ZIP_PATH
from io import BytesIO
import csv
from io import TextIOWrapper


def test_check_pdf():    # тут начинается pdf
    with zipfile.ZipFile(ZIP_PATH, 'r', zipfile.ZIP_DEFLATED) as zp:
        with zp.open('Python Testing with Pytest (Brian Okken).pdf', 'r') as file:
            reader = PdfReader(file)
            total_count = len(reader.pages)
            keyword = 'Ruiz'
            assert any(keyword in page.extract_text() for page in reader.pages), f"Keyword '{keyword}' not found in PDF"

        assert total_count == 256

def test_check_xlsx():    # тут начинается xlsx
    zip_ = 'resource/test.zip'
    with zipfile.ZipFile(zip_, 'r', zipfile.ZIP_DEFLATED) as zip_file:
        with zip_file.open('file_example_XLSX_50.xlsx') as zip_xlsx:
            workbook = load_workbook(zip_xlsx)
            sheet = workbook.active
            wb = openpyxl.load_workbook(zip_xlsx)
            length = len(wb.sheetnames)
            sheet_xlsx = wb.sheetnames
            row_count = sheet.max_row
            column_count = sheet.max_column
            value_1 = sheet.cell(row=51, column=8).value
            value_2 = sheet.cell(row=2, column=2).value

    assert length == 1
    assert sheet_xlsx == ['Sheet1']
    assert row_count == 51
    assert column_count == 8
    assert value_1 == 6125
    assert value_2 == 'Dulce'

def test_check_xls():
        # тут начинается xls
    with zipfile.ZipFile(ZIP_PATH, 'r', zipfile.ZIP_DEFLATED) as zip_file:
        with zip_file.open('file_example_XLS_10.xls') as zip_xls:
            content = zip_xls.read()
            # Создаем буферный объект BytesIO для содержимого файла
            buffer = BytesIO(content)
            # Открываем файл с использованием буфера
            workbook1 = open_workbook(file_contents=buffer.read())
            count_xls = workbook1.nsheets
            name_xls = workbook1.sheet_names()
            sheet = workbook1.sheet_by_index(0)
            count_rows = sheet.nrows
            count_col = sheet.ncols
            value_3 = sheet.cell_value(rowx=5, colx=1)

    assert count_xls == 1
    assert name_xls == ['Sheet1']
    assert count_rows == 10
    assert count_col == 8
    assert value_3 == 'Nereida'

def test_check_csv():
        # тут начинается csv
    with zipfile.ZipFile(ZIP_PATH, 'r', zipfile.ZIP_DEFLATED) as zip_file:
        with zip_file.open('file_example_csv_10.csv') as zip_csv:
            reader = list(csv.reader(TextIOWrapper(zip_csv, 'utf-8-sig'), delimiter=';'))
            nine_string = reader[9]
            cross_cell = reader[5]


    assert nine_string[1] == 'Vincenza'
    assert cross_cell[4] == 'United States'
